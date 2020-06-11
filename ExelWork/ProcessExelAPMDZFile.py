#!/usr/bin/env python
# -*- coding: utf-8 -*-

import openpyxl
import codecs
import re
import os.path


class ReadingFromExelFile_APMDZ_SNs(object):
    '''
    '''

    def __init__(self, Full_path_to__ExelFile):
        '''
        '''

        if os.path.exists(Full_path_to__ExelFile):
            self.__FullPathToExelFile = Full_path_to__ExelFile
            wb = openpyxl.load_workbook(self.__FullPathToExelFile)
            sp = wb.get_sheet_names()
            self.__Sheet = wb.get_sheet_by_name(sp[0])
            self.__RezultList = list()
            self.ConstructList()

        else:
            raise IOError("File Not Exist")




    def ConstructList(self):
        '''

        :return:
        '''
        for num in range(2, 1205):
            v = self.__Sheet.cell(row=num, column=2).value
        #print(v)
        #print(type(v))
            self.Adding__APMDZ_SNs(v)



    def Adding__APMDZ_SNs(self, SN):
        '''
        :param SN:
        :return:
        '''
        s = ''
        if type(SN) is int:
            ts = str(SN);
            s = "0000" + ts
            #spam += s
        else:
            s = SN
        self.__RezultList.append(s)
            #spam += v
        #spam += ', №'


    def RemoveSNs(self,spisok):
        '''
        :param spisok:
        :return:
        '''
        for e in self.__RezultList:
            for s in spisok:
                if s == e:
                    self.__RezultList.remove(s)

    def Count(self):
        '''

        :return:
        '''
        return len(self.__RezultList)


    def WriteAll__APMDZ_SNs(self, outputfile):
        '''
        :param outputfile:
        :return:
        '''
        res_string = ""
        for i,s in enumerate(self.__RezultList):
            res_string = res_string + '№'
            res_string = res_string + s
            res_string = res_string + ', '
        with codecs.open(outputfile, "w", "cp1251") as g:
            g.write(res_string)

    def Reconstruct__APMDZ_SNsFromFile(self,infile):
        '''

        :param infile:
        :return:
        '''
        if os.path.exists(infile):
            with codecs.open(infile, "r", "cp1251") as g:
                s = g.read()
                self.__RezultList = re.split(r", №", s)
                for i,s in enumerate(self.__RezultList):
                    for c in s:
                        if c == "№" or c == " " or c == "," :
                            print(s)
                            pPattern = re.compile( r"(?:\W\w)*(\d\d\d\d\d\d\d\d)(?:\W\w)*", re.S | re.L)
                            M1 = pPattern.search(s)
                            if M1:
                                print (M1.group(1))
                                #(headerString, number) = pPattern.subn(M1.group(1) + MACRO, headerString, 1)
                                s = M1.group(1)
                                self.__RezultList[i] = s
        else:
            raise IOError("File Not Exist")
        return 0


#print(sp[0])
#print(sheet.cell(row=2, column=2).value)
#print(sheet.cell(row=3, column=2).value)
#print(wb.get_named_ranges())
# print (sheet.columns[1])
# for num in sheet.columns[1] :
#    print(num.value)

#file_path = u"e:\\spesok_2.txt"

# if file_path:
#    with codecs.open(file_path, "w", "cp1251") as g:
#        g.write(self.__SourceString)
# else:
#    with codecs.open(self.__FullPath, "w", "cp1251") as g:
#        g.write(self.__SourceString)

#spam = ""
#for num in range(2, 1205):
#    v = sheet.cell(row=num, column=2).value
#    print(v)
#    print(type(v))
#    if type(v) is int:
#        ts = str(v);
#        s = "0000" + ts
#        spam += s
#    else:
#        spam += v
#    spam += ', №'
#with codecs.open(file_path, "w", "cp1251") as g:
#    g.write(spam)





# print(sheet.get_named_ranges())
# for num in sheet.columns[2] :





# print(sheet.get_highest_row())
if __name__ == "__main__":
     inFile = 'e:\\Rabota\\PythonProjects\\ExelWork\\spisok.xlsx'
     outFile_1 = 'e:\\Rabota\\PythonProjects\\ExelWork\\spisok_1.txt'
     cReadingFromExelFile_APMDZ_SNs = ReadingFromExelFile_APMDZ_SNs(inFile)
     cReadingFromExelFile_APMDZ_SNs.WriteAll__APMDZ_SNs(outFile_1)

     print (cReadingFromExelFile_APMDZ_SNs.Count())

     spisok = ["00001308", "00002453", "00002811", "00002878", "00003504", "00003681", "00003842", "00004287"]
     cReadingFromExelFile_APMDZ_SNs.RemoveSNs(spisok)
     print(cReadingFromExelFile_APMDZ_SNs.Count())

     outFile_2 = 'e:\\Rabota\\PythonProjects\\ExelWork\\spisok_2.txt'
     cReadingFromExelFile_APMDZ_SNs.WriteAll__APMDZ_SNs(outFile_2)

     cReadingFromExelFile_APMDZ_SNs.Reconstruct__APMDZ_SNsFromFile(outFile_2)
     outFile_3 = 'e:\\Rabota\\PythonProjects\\ExelWork\\spisok_3.txt'
     cReadingFromExelFile_APMDZ_SNs.WriteAll__APMDZ_SNs(outFile_3)

     print(cReadingFromExelFile_APMDZ_SNs.Count())




