#!/usr/bin/env python
# -*- coding: utf-8 -*-

import openpyxl
import codecs
import re
import os.path


wb = openpyxl.load_workbook('e:\\Rabota\\PythonProjects\\ExelWork\\spisok.xlsx')

sp = wb.get_sheet_names()
sheet = wb.get_sheet_by_name(sp[0])
print(sp[0])
print(sheet.cell(row=2, column=2).value)
print(sheet.cell(row=3, column=2).value)
print (wb.get_named_ranges())
#print (sheet.columns[1])
#for num in sheet.columns[1] :
#    print(num.value)

file_path = u"e:\\spesok_2.txt"

#if file_path:
#    with codecs.open(file_path, "w", "cp1251") as g:
#        g.write(self.__SourceString)
#else:
#    with codecs.open(self.__FullPath, "w", "cp1251") as g:
#        g.write(self.__SourceString)

spam = ""
for num in range(2,1205) :
    v = sheet.cell(row=num, column=2).value
    print(v)
    print(type(v))
    if type(v) is int :
        ts=str(v);
        s = "0000" + ts
        spam += s
    else :
        spam += v
    spam += ', №'
with codecs.open(file_path, "w", "cp1251") as g:
        g.write(spam)





#print(sheet.get_named_ranges())
#for num in sheet.columns[2] :





#print(sheet.get_highest_row())
#if __name__ == "__main__":
